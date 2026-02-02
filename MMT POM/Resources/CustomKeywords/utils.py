from robot.api.deco import keyword
from openpyxl import load_workbook
import random
import datetime

@keyword
def fetch_testdata_by_id(file_path, target_id):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if col_name and ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                workbook.close()
                return data_dict

        workbook.close()
        raise ValueError(f"Target ID '{target_id}' not found in the Excel file.")
    
    except Exception as e:
        if 'workbook' in locals():
            workbook.close()
        print(f"Error reading Excel file: {e}")
        raise

@keyword
def fetch_all_testdata(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        all_data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                data_dict = {}
                for col_name, value in zip(header, row):
                    if col_name and ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                all_data.append(data_dict)
        
        workbook.close()
        return all_data
    
    except Exception as e:
        if 'workbook' in locals():
            workbook.close()
        print(f"Error reading Excel file: {e}")
        raise

@keyword
def generate_random_pickup_time():
    try:
        # Generate random hour (1-12)
        hour = random.randint(1, 12)
        
        # Generate random minute from allowed 5-minute intervals
        allowed_minutes = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]
        minute = random.choice(allowed_minutes)
        
        # Generate random meridian (AM/PM)
        meridian = random.choice(['AM', 'PM'])
        
        # Format time as "H:MM AM/PM"
        formatted_time = f"{hour}:{minute:02d} {meridian}"
        
        print(f"Generated random pickup time: {formatted_time}")
        return formatted_time
    
    except Exception as e:
        print(f"Error generating random pickup time: {e}")
        raise

@keyword
def generate_random_departure_date():
    try:
        # Get today's date
        today = datetime.date.today()
        
        # Calculate 6 months from today (approximately 180 days)
        six_months_later = today + datetime.timedelta(days=180)
        
        # Generate random number of days between 0 and 180
        random_days = random.randint(0, 180)
        
        # Calculate random date
        random_date = today + datetime.timedelta(days=random_days)
        
        # Convert to datetime object (required for Robot Framework date processing)
        random_datetime = datetime.datetime.combine(random_date, datetime.time())
        
        print(f"Generated random departure date: {random_datetime.strftime('%Y-%m-%d')}")
        return random_datetime
    
    except Exception as e:
        print(f"Error generating random departure date: {e}")
        raise