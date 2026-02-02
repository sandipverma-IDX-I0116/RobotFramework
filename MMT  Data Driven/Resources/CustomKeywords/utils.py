import openpyxl
from robot.api.deco import keyword
import random
from datetime import date, timedelta, datetime as dt


@keyword("Fetch Testdata By Id")
def fetch_testdata_by_id(excel_file_path, sheet_name, tc_id):
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        # tc_id_column_index = 1
        tc_id_column_index = None

        for idx, header in enumerate(headers, start=1):
            if header and str(header).upper() in ['TC_ID', 'TEST_ID', '${TC_ID}', '${TEST_ID}']:
                tc_id_column_index = idx
                break
        if tc_id_column_index is None:
            raise ValueError(f"TC_ID column not found in Excel headers: {headers}")
        test_data = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[tc_id_column_index - 1].value == tc_id:
                for col_idx, header in enumerate(headers, start=1):
                    if header:
                        clean_header = str(header).replace('${', '').replace('}', '').strip()
                        test_data[clean_header] = row[col_idx - 1].value
                break
        if not test_data:
            raise ValueError(f"Test Case ID '{tc_id}' not found in Excel file")
        wb.close()
        return test_data
    except Exception as e:
        raise Exception(f"Error fetching test data for TC_ID '{tc_id}': {str(e)}")


@keyword("Create Dict With Dot Access")
def create_dict_with_dot_access(data_dict):
    """Convert dictionary to object with dot notation access for Robot Framework"""
    class DotDict:
        def __init__(self, d):
            for key, value in d.items():
                setattr(self, key, value)
        def __getitem__(self, key):
            return getattr(self, key)
        def __str__(self):
            return str(self.__dict__)
    return DotDict(data_dict)

@keyword
def generate_random_pickup_time():
    try:
        # Generate random hour (1-12)
        hour = random.randint(1, 12)
        
        # Generate random minute from allowed 5-minute intervals
        allowed_minutes = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]
        minute = random.choice(allowed_minutes)
        
        # Generate random meridian (AM/PM)
        meridian = random.choice(['AM', 'PM'])
        
        # Format time as "H:MM AM/PM"
        formatted_time = f"{hour:02d}:{minute:02d} {meridian}"
        
        print(f"Generated random pickup time: {formatted_time}")
        return formatted_time
    
    except Exception as e:
        print(f"Error generating random pickup time: {e}")
        raise

@keyword
def generate_random_departure_date():
    try:
        today = date.today()
        
        six_months_later = today + timedelta(days=180)
        
        random_days = random.randint(0, 180)
        
        random_date = today + timedelta(days=random_days)
        
        random_datetime = dt.combine(random_date, dt.min.time())
        
        print(f"Generated random departure date: {random_datetime.strftime('%Y-%m-%d')}")
        return random_datetime
    
    except Exception as e:
        print(f"Error generating random departure date: {e}")
        raise